# -*- coding: utf-8 -*-
"""
Cable Calculation — v11.0
Six-tab Tkinter GUI matching your flow:
  1) Project Setup
  2) Load & System
  3) Installation & Environment
  4) Overload Protection Check
  5) Voltage Drop
  6) Summary & Export
"""
from __future__ import annotations

import os, sys, math, json, glob, re, datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont
from typing import Optional, Dict, List, Tuple

APP_VERSION = "v11.0"
PROJECT_DIR = os.path.join(os.path.expanduser("~"), "CableCalcProjects")
os.makedirs(PROJECT_DIR, exist_ok=True)
RECENTS_FILE = os.path.join(PROJECT_DIR, "_recent_projects.json")

HAVE_XLSX = True
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    HAVE_XLSX = False

# --- Free-Air grouping tables (sample values you provided) ---
FREE_AIR_MC = {
    "Horizontal Perforated Cable Tray (Touching)": {
        1: {1:1.00, 2:0.88, 3:0.82, 4:0.79, 6:0.76, 9:0.73},
        2: {1:1.00, 2:0.87, 3:0.80, 4:0.77, 6:0.73, 9:0.68},
        3: {1:1.00, 2:0.86, 3:0.79, 4:0.76, 6:0.71, 9:0.66},
        6: {1:1.00, 2:1.00, 3:0.77, 4:0.73, 6:0.68, 9:0.64},
    },
    "Horizontal Perforated Cable Tray (Spaced)": {
        1: {1:1.00, 2:1.00, 3:0.98, 4:0.95, 6:0.91},
        2: {1:1.00, 2:0.99, 3:0.96, 4:0.92, 6:0.87},
        3: {1:1.00, 2:0.98, 3:0.95, 4:0.91, 6:0.85},
    },
    "Vertical Perforated Tray (Touching)": {
        1: {1:1.00, 2:0.88, 3:0.82, 4:0.78, 6:0.73, 9:0.72},
        2: {1:1.00, 2:0.88, 3:0.81, 4:0.76, 6:0.71, 9:0.70},
    },
    "Vertical Perforated Tray (Spaced)": {
        1: {1:1.00, 2:0.91, 3:0.89, 4:0.88, 6:0.87},
        2: {1:1.00, 2:0.91, 3:0.88, 4:0.87, 6:0.85},
    },
    "Unperforated Horizontal Cable Tray (Touching)": {
        1: {1:0.97, 2:0.84, 3:0.78, 4:0.75, 6:0.71, 9:0.68},
        2: {1:0.97, 2:0.83, 3:0.76, 4:0.72, 6:0.68, 9:0.63},
        3: {1:0.97, 2:0.82, 3:0.75, 4:0.71, 6:0.66, 9:0.61},
        6: {1:0.97, 2:0.81, 3:0.73, 4:0.69, 6:0.63, 9:0.58},
    },
    "Cable Ladder System (Touching)": {
        1: {1:1.00, 2:0.87, 3:0.82, 4:0.80, 6:0.79, 9:0.78},
        2: {1:1.00, 2:0.86, 3:0.80, 4:0.78, 6:0.76, 9:0.73},
        3: {1:1.00, 2:0.85, 3:0.79, 4:0.76, 6:0.73, 9:0.70},
        6: {1:1.00, 2:0.84, 3:0.77, 4:0.73, 6:0.68, 9:0.64},
    },
    "Cable Ladder System (Spaced)": {
        1: {1:1.00, 2:1.00, 3:1.00, 4:1.00, 6:1.00},
        2: {1:1.00, 2:0.99, 3:0.98, 4:0.97, 6:0.96},
        3: {1:1.00, 2:0.98, 3:0.97, 4:0.96, 6:0.93},
    },
}
SC_CABLE_MGMT_OPTIONS = [
    "Horizontal Perforated Cable (Spaced)",
    "Vertical Perforated Tray (Spaced)",
    "Horizontal Ladder System (Spaced)",
]
FREE_AIR_SC_TREFOIL = {
    "Horizontal Perforated Cable (Spaced)": {1: {1:1.00, 2:0.98, 3:0.96}, 2: {1:0.97, 2:0.93, 3:0.89}, 3: {1:0.96, 2:0.92, 3:0.86}},
    "Vertical Perforated Tray (Spaced)":    {1: {1:1.00, 2:0.91, 3:0.89}, 2: {1:1.00, 2:0.90, 3:0.86}},
    "Horizontal Ladder System (Spaced)":    {1: {1:1.00, 2:1.00, 3:1.00}, 2: {1:0.97, 2:0.95, 3:0.93}, 3: {1:0.96, 2:0.94, 3:0.90}},
}
FREE_AIR_SC_FLAT = {
    "Horizontal Perforated Cable (Spaced)": {1: {1:0.98, 2:0.91, 3:0.87}, 2: {1:0.96, 2:0.87, 3:0.81}, 3: {1:0.95, 2:0.85, 3:0.78}},
    "Vertical Perforated Tray (Spaced)":    {1: {1:0.96, 2:0.86},          2: {1:0.95, 2:0.84}},
}
TEMP_AIR_FACTORS = {
    "PVC":  {25: 1.49, 30: 1.40, 35: 1.31, 40: 1.22, 45: 1.12, 50: 1.00, 55: 0.87},
    "XLPE": {25: 1.28, 30: 1.23, 35: 1.18, 40: 1.13, 45: 1.06, 50: 1.00, 55: 0.94},
}
PROTECTION_I2_MULT = {
    "MCB (IEC 60898-1)": 1.45,
    "MCCB (IEC 60947-2)": 1.30,
    "Fuse gG (IEC 60269)": 1.60,
    "Fuse aM": 1.60,
    "Custom": None,
}

# --- helpers ---
def _read_json(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _write_json(path: str, data: dict) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

def _safe_filename(name: str) -> str:
    return re.sub(r"[^\w\-]+", "_", (name or "").strip())[:80] or "Project"

def project_filepath(project_no: str) -> str:
    return os.path.join(PROJECT_DIR, f"{_safe_filename(project_no or 'UNSPECIFIED')}.xlsx")

def add_to_recents(path: str) -> None:
    data = _read_json(RECENTS_FILE)
    items = [p for p in data.get("items", []) if p != path]
    items.insert(0, path)
    data["items"] = items[:12]
    _write_json(RECENTS_FILE, data)

def get_recents() -> List[str]:
    return _read_json(RECENTS_FILE).get("items", [])

# --- numeric calcs ---
def calc_phase_voltage(v_ll: float, phase: str) -> float:
    return round(v_ll / math.sqrt(3.0), 2) if phase == "3" else round(v_ll, 2)

def calc_kva_kw_pf_eff(kw: float, pf: float, eff_pct: float) -> float:
    eff = max(0.01, eff_pct / 100.0)
    return round(kw / (pf * eff), 3) if kw > 0 and pf > 0 else 0.0

def calc_flc(kw: float, v_ll: float, phase: str, pf: float, eff_pct: float) -> float:
    eff = max(0.01, eff_pct / 100.0)
    if kw <= 0 or v_ll <= 0 or pf <= 0 or eff <= 0:
        return 0.0
    p = kw * 1000.0
    i = p / (math.sqrt(3.0) * v_ll * pf * eff) if phase == "3" else p / (v_ll * pf * eff)
    return round(i, 2)

def calc_start_current(flc: float, start_factor: float) -> float:
    return round(max(0.0, flc * start_factor), 2)

# --- Excel Project store ---
class ProjectStore:
    def __init__(self, path: str):
        self.path = path
        self.wb = None

    def open_or_create(self):
        if not HAVE_XLSX:
            raise RuntimeError("openpyxl not available")
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            if "Project" not in self.wb.sheetnames:
                self.wb.create_sheet("Project")
            if "Calc" not in self.wb.sheetnames:
                ws_calc = self.wb.create_sheet("Calc")
                ws_calc.append(self._calc_headers())
            else:
                ws_calc = self.wb["Calc"]
                if ws_calc.max_row < 1:
                    ws_calc.append(self._calc_headers())
                else:
                    self._ensure_calc_headers(ws_calc)
        else:
            self.wb = Workbook()
            ws_proj = self.wb.active
            ws_proj.title = "Project"
            ws_calc = self.wb.create_sheet("Calc")
            ws_calc.append(self._calc_headers())
        self.save()

    def _calc_headers(self) -> List[str]:
        return [
            "CalculationNumber","FromDescription","ToDescription",
            "FromTag","ToTag","CircuitID","CableTag",
            "Phase","SystemVoltageLL_V","PhaseVoltage_V",
            "EquipmentRating_kW","PowerFactor","Efficiency_pct",
            "StartFactor","StartPF","ApparentPower_kVA",
            "FullLoadCurrent_A","StartingCurrent_A",
            "ProtectiveDevice_A","OverloadSetting_pct","CircuitLength_m",
            "TypeOfCable","TypeOfInstallation","CoreType","Formation",
            "DepthOfCable","DepthCF","SoilThermalResistivity","SoilResistivityCF",
            "NumberOfCircuits","GroundOrDuctsCF","SpacingBetweenCircuits",
            "CableManagement","NumberOfTraysLadders","NumberOfCablesPerTray",
            "InAirDeratingFactor","AmbientTemperature_C","TemperatureDeratingFactor",
            "OverallDeratingFactor",
            "ProtectionType","I2_Factor","CableRatedCurrent_Iz_single",
            "CablesInParallel_n","TotalCableRating_Iz","OverloadCheck_OK",
            "Cable_r_ohm_per_km","Cable_x_ohm_per_km",
            "SteadyVD_V","SteadyVD_pct","SteadyVD_limit_pct","SteadyVD_OK",
            "StartingVD_V","StartingVD_pct","StartingVD_limit_pct","StartingVD_OK",
            "SavedAt",
        ]

    def _ensure_calc_headers(self, ws_calc):
        existing = [c.value for c in ws_calc[1]]
        want = self._calc_headers()
        to_add = [h for h in want if h not in existing]
        if to_add:
            for idx, h in enumerate(to_add, start=len(existing)+1):
                ws_calc.cell(1, idx).value = h

    def save(self):
        if self.wb:
            self.wb.save(self.path)

    # Project KV
    def write_project_info(self, kv: Dict[str,str]):
        ws = self.wb["Project"]
        ws.delete_rows(1, ws.max_row)
        for k, v in kv.items():
            ws.append([k, v])
        self.save()

    def read_project_info(self) -> Dict[str,str]:
        ws = self.wb["Project"]
        out = {}
        for r in ws.iter_rows(min_row=1, values_only=True):
            if r and r[0] is not None:
                out[str(r[0])] = "" if len(r) < 2 or r[1] is None else str(r[1])
        return out

    # Calc helpers
    def list_calc_numbers(self) -> List[str]:
        ws = self.wb["Calc"]
        nums = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and str(v).strip() != "":
                nums.append(str(v))
        return nums

    def read_calc_by_index(self, idx: int) -> Optional[Dict[str, object]]:
        ws = self.wb["Calc"]
        if ws.max_row < 2: return None
        headers = [c.value for c in ws[1]]
        rownum = 2 + idx
        if rownum < 2 or rownum > ws.max_row: return None
        row = [ws.cell(rownum, c).value for c in range(1, len(headers)+1)]
        return dict(zip(headers, row))

    def read_calc_by_number(self, calcno: str) -> Optional[Tuple[int, Dict[str, object]]]:
        ws = self.wb["Calc"]
        headers = [c.value for c in ws[1]]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "") == str(calcno):
                row = [ws.cell(r, c).value for c in range(1, len(headers)+1)]
                return (r - 2, dict(zip(headers, row)))
        return None

    def search_calculations(self, query: str):
        q = (query or "").strip().lower()
        if not q: return []
        ws = self.wb["Calc"]
        if ws.max_row < 2: return []
        headers = [c.value for c in ws[1]]
        results = []
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            hay = " ".join("" if v is None else str(v) for v in row_vals).lower()
            if q in hay:
                results.append((r - 2, dict(zip(headers, row_vals))))
        return results

    def upsert_calc_row(self, d: Dict[str, object]):
        ws = self.wb["Calc"]
        self._ensure_calc_headers(ws)
        headers = [c.value for c in ws[1]]

        key = str(d.get("CalculationNumber",""))
        if not key:
            raise ValueError("Calculation Number required")

        target = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "") == key:
                target = r
                break
        if target is None:
            target = ws.max_row + 1

        for name, value in d.items():
            try:
                col = headers.index(name) + 1
            except ValueError:
                col = len(headers) + 1
                ws.cell(1, col).value = name
                headers.append(name)
            ws.cell(target, col).value = value

        try:
            col = headers.index("SavedAt") + 1
        except ValueError:
            col = len(headers) + 1
            ws.cell(1, col).value = "SavedAt"
            headers.append("SavedAt")
        ws.cell(target, col).value = datetime.datetime.now().isoformat(timespec="seconds")
        self.save()
class CableCalcApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"Cable Calculation — {APP_VERSION}")
        self.geometry("1320x780")
        self.minsize(1220, 720)

        self.BG  = "#f5f5dc"
        self.EBG = "#fffaf0"
        self.configure(bg=self.BG)

        style = ttk.Style(self)
        try: style.theme_use("clam")
        except Exception: pass
        style.configure(".", background=self.BG)
        style.configure("App.TFrame", background=self.BG)
        style.configure("App.TLabel", background=self.BG)
        style.configure("App.Section.TLabel", background=self.BG, font=("Segoe UI", 10, "bold"))
        style.configure("App.TButton", padding=4)
        style.configure("App.TEntry", fieldbackground=self.EBG)
        style.configure("TNotebook", background=self.BG)
        style.configure("TNotebook.Tab", padding=(12, 4))

        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.store: Optional[ProjectStore] = None
        self.dirty_project = False
        self.dirty_calc = False
        self._autosave_job = None
        self._autosave_ms = 600

        self.box_font: Optional[tkfont.Font] = None
        self.calc_order: List[str] = []
        self.calc_index: int = -1
        self._tab2_boot_loaded: bool = False
        self.cm_label_map: Dict[str, str] = {}

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=2, pady=2)
        self.tab1 = ttk.Frame(self.nb, style="App.TFrame")
        self.tab2 = ttk.Frame(self.nb, style="App.TFrame")
        self.tab3 = ttk.Frame(self.nb, style="App.TFrame")
        self.tab4 = ttk.Frame(self.nb, style="App.TFrame")
        self.tab5 = ttk.Frame(self.nb, style="App.TFrame")
        self.tab6 = ttk.Frame(self.nb, style="App.TFrame")

        self.nb.add(self.tab1, text="Project Setup")
        self.nb.add(self.tab2, text="Load & System")
        self.nb.add(self.tab3, text="Installation & Env")
        self.nb.add(self.tab4, text="Overload Check")
        self.nb.add(self.tab5, text="Voltage Drop")
        self.nb.add(self.tab6, text="Summary & Export")

        self._build_tab1()
        self._build_tab2()
        self._build_tab3()
        self._build_tab4()
        self._build_tab5()
        self._build_tab6()

        self._mirror_to_tab2(self._collect_project_info())
        self._mirror_to_tab3_project(self._collect_project_info())
        self._mirror_to_tab4_project(self._collect_project_info())
        self._mirror_to_tab5_project(self._collect_project_info())
        self._mirror_to_tab6_project(self._collect_project_info())

        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    # --- utilities ---
    def _install_text_bindings(self, widget: tk.Widget):
        widget.bind("<Control-c>", lambda e: (widget.event_generate("<<Copy>>"), "break"))
        widget.bind("<Control-x>", lambda e: (widget.event_generate("<<Cut>>"), "break"))
        widget.bind("<Control-v>", lambda e: (widget.event_generate("<<Paste>>"), "break"))
        widget.bind("<Control-a>", lambda e: (self._select_all(widget), "break"))

    def _select_all(self, widget: tk.Widget):
        try:
            widget.select_range(0, "end"); widget.icursor("end")
        except Exception:
            try: widget.tag_add("sel", "1.0", "end-1c")
            except Exception: pass

    def _schedule_autosave(self):
        if self._autosave_job is not None:
            self.after_cancel(self._autosave_job)
        self._autosave_job = self.after(self._autosave_ms, self._attempt_autosave_current_calc)

    def _attempt_autosave_current_calc(self):
        self._autosave_job = None
        if self.store is None: return
        calcno = (getattr(self, "t2_calcno", None).get() if getattr(self, "t2_calcno", None) else "")
        calcno = (calcno or "").strip()
        if not calcno: return
        try: self.on_save_calc()
        except Exception: pass

    def _set_ro(self, entry: ttk.Entry, text: str):
        entry.configure(state="normal"); entry.delete(0, "end")
        entry.insert(0, text); entry.configure(state="readonly")

    def _to_float(self, s, default=0.0):
        try: return float(str(s).strip())
        except Exception: return default
    # ---------- Tab 1 (Project Setup) ----------
    def _build_tab1(self):
        f = self.tab1
        pad = {"padx": 10, "pady": 5}
        ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=8, sticky="w", **pad)

        ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad)
        self.t1_title = ttk.Entry(f, width=50, style="App.TEntry"); self.t1_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
        ttk.Label(f, text="Project Number (required)", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad)
        self.t1_number = ttk.Entry(f, width=25, style="App.TEntry"); self.t1_number.grid(row=1, column=4, sticky="we", **pad)
        ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad)
        self.t1_location = ttk.Entry(f, width=25, style="App.TEntry"); self.t1_location.grid(row=1, column=6, sticky="we", **pad)

        ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad)
        self.t1_docno = ttk.Entry(f, width=50, style="App.TEntry"); self.t1_docno.grid(row=2, column=1, sticky="we", **pad)
        ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad)
        self.t1_rev = ttk.Entry(f, width=10, style="App.TEntry"); self.t1_rev.grid(row=2, column=4, sticky="we", **pad)

        ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=8, sticky="we", padx=8, pady=6)
        ttk.Label(f, text="Project Description", style="App.TLabel").grid(row=4, column=0, sticky="nw", **pad)
        self.t1_desc = tk.Text(f, width=110, height=8, bg=self.EBG, relief="sunken"); self.t1_desc.grid(row=4, column=1, columnspan=6, sticky="nsew", **pad)
        ttk.Separator(f, orient="horizontal").grid(row=5, column=0, columnspan=8, sticky="we", padx=8, pady=6)
        ttk.Label(f, text="Project Notes", style="App.TLabel").grid(row=6, column=0, sticky="nw", **pad)
        self.t1_notes = tk.Text(f, width=110, height=8, bg=self.EBG, relief="sunken"); self.t1_notes.grid(row=6, column=1, columnspan=6, sticky="nsew", **pad)

        for c in (1,6): f.grid_columnconfigure(c, weight=1)
        f.grid_rowconfigure(4, weight=1); f.grid_rowconfigure(6, weight=1)
        ttk.Separator(f, orient="horizontal").grid(row=7, column=0, columnspan=8, sticky="we", padx=8, pady=6)

        ttk.Button(f, text="Create New Project", style="App.TButton", command=self.on_create_project).grid(row=8, column=0, sticky="w", **pad)
        ttk.Button(f, text="Open Recent...", style="App.TButton", command=self.on_open_recent).grid(row=8, column=1, sticky="w", **pad)
        ttk.Button(f, text="Update Save Project", style="App.TButton", command=self.on_save_project).grid(row=8, column=3, sticky="w", **pad)
        self.t1_search = ttk.Entry(f, width=36, style="App.TEntry"); self.t1_search.grid(row=8, column=4, sticky="we", **pad)
        ttk.Button(f, text="Search Project", style="App.TButton", command=self.on_search_project).grid(row=8, column=5, sticky="w", **pad)
        ttk.Button(f, text="Next", style="App.TButton", command=lambda: self.nb.select(self.tab2)).grid(row=8, column=6, sticky="e", **pad)

        self.box_font = tkfont.Font(font=self.t1_title.cget("font"))
        for w in [self.t1_title, self.t1_number, self.t1_location, self.t1_docno, self.t1_rev, self.t1_search]:
            w.configure(font=self.box_font); self._install_text_bindings(w)
        self.t1_desc.configure(font=self.box_font); self.t1_notes.configure(font=self.box_font)
        self._install_text_bindings(self.t1_desc); self._install_text_bindings(self.t1_notes)

        for w in [self.t1_title, self.t1_number, self.t1_location, self.t1_docno, self.t1_rev]:
            w.bind("<KeyRelease>", lambda e: self._mark_dirty_project())
        for t in [self.t1_desc, self.t1_notes]:
            t.bind("<<Modified>>", self._on_text1_modified)

    def _on_text1_modified(self, e):
        w = e.widget
        if w.edit_modified():
            w.edit_modified(False)
            self._mark_dirty_project()

    def _mark_dirty_project(self):
        self.dirty_project = True

    def _collect_project_info(self) -> Dict[str, str]:
        def val(entry):
            try: return entry.get().strip()
            except Exception: return ""
        def text_val(txt: tk.Text) -> str:
            try: return txt.get("1.0", "end-1c").strip()
            except Exception: return ""
        return {
            "Project Title":        val(getattr(self, "t1_title", None)),
            "Project Number":       val(getattr(self, "t1_number", None)),
            "Location":             val(getattr(self, "t1_location", None)),
            "Document Number":      val(getattr(self, "t1_docno", None)),
            "Rev No":               val(getattr(self, "t1_rev", None)),
            "Project Description":  text_val(getattr(self, "t1_desc", None)),
            "Project Notes":        text_val(getattr(self, "t1_notes", None)),
        }

    def _fill_tab1_from_kv(self, kv: Dict[str, str]) -> None:
        def set_entry(entry, text):
            if not entry: return
            try: entry.delete(0, "end"); entry.insert(0, text or "")
            except Exception: pass
        def set_text(txt: tk.Text, text: str):
            if not txt: return
            try: txt.delete("1.0", "end"); txt.insert("1.0", text or "")
            except Exception: pass
        set_entry(getattr(self, "t1_title", None),    kv.get("Project Title", ""))
        set_entry(getattr(self, "t1_number", None),   kv.get("Project Number", ""))
        set_entry(getattr(self, "t1_location", None), kv.get("Location", ""))
        set_entry(getattr(self, "t1_docno", None),    kv.get("Document Number", ""))
        set_entry(getattr(self, "t1_rev", None),      kv.get("Rev No", ""))
        set_text(getattr(self, "t1_desc", None),  kv.get("Project Description", ""))
        set_text(getattr(self, "t1_notes", None), kv.get("Project Notes", ""))

    # --- project mirrors (Tab 2..6 headers) ---
    def _mirror_to_tab2(self, kv: Dict[str,str]):
        def ro(entry: ttk.Entry, val: str):
            entry.configure(state="normal"); entry.delete(0,"end")
            entry.insert(0, val or ""); entry.configure(state="readonly")
        ro(self.t2_title,   kv.get("Project Title",""))
        ro(self.t2_number,  kv.get("Project Number",""))
        ro(self.t2_location,kv.get("Location",""))
        ro(self.t2_docno,   kv.get("Document Number",""))
        ro(self.t2_rev,     kv.get("Rev No",""))

    def _mirror_to_tab3_project(self, kv: Dict[str,str]):
        def ro(entry: ttk.Entry, val: str):
            entry.configure(state="normal"); entry.delete(0,"end")
            entry.insert(0, val or ""); entry.configure(state="readonly")
        ro(self.t3_title,   kv.get("Project Title",""))
        ro(self.t3_number,  kv.get("Project Number",""))
        ro(self.t3_location,kv.get("Location",""))
        ro(self.t3_docno,   kv.get("Document Number",""))
        ro(self.t3_rev,     kv.get("Rev No",""))

    def _mirror_to_tab4_project(self, kv: Dict[str,str]):
        def ro(entry: ttk.Entry, val: str):
            entry.configure(state="normal"); entry.delete(0,"end")
            entry.insert(0, val or ""); entry.configure(state="readonly")
        ro(self.t4_title,   kv.get("Project Title",""))
        ro(self.t4_number,  kv.get("Project Number",""))
        ro(self.t4_location,kv.get("Location",""))
        ro(self.t4_docno,   kv.get("Document Number",""))
        ro(self.t4_rev,     kv.get("Rev No",""))

    def _mirror_to_tab5_project(self, kv: Dict[str,str]):
        def ro(entry: ttk.Entry, val: str):
            entry.configure(state="normal"); entry.delete(0,"end")
            entry.insert(0, val or ""); entry.configure(state="readonly")
        ro(self.t5_title,   kv.get("Project Title",""))
        ro(self.t5_number,  kv.get("Project Number",""))
        ro(self.t5_location,kv.get("Location",""))
        ro(self.t5_docno,   kv.get("Document Number",""))
        ro(self.t5_rev,     kv.get("Rev No",""))

    def _mirror_to_tab6_project(self, kv: Dict[str,str]):
        def ro(entry: ttk.Entry, val: str):
            entry.configure(state="normal"); entry.delete(0,"end")
            entry.insert(0, val or ""); entry.configure(state="readonly")
        ro(self.t6_title,   kv.get("Project Title",""))
        ro(self.t6_number,  kv.get("Project Number",""))
        ro(self.t6_location,kv.get("Location",""))
        ro(self.t6_docno,   kv.get("Document Number",""))
        ro(self.t6_rev,     kv.get("Rev No",""))

    # --- Project load/save/open/search ---
    def _load_project(self, path: str):
        try:
            if not HAVE_XLSX:
                messagebox.showerror("Excel backend", "openpyxl is not available."); return
            if not path: return
            self.store = ProjectStore(path); self.store.open_or_create()
            kv = self.store.read_project_info()
            self._fill_tab1_from_kv(kv)
            self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv)
            self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
            try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(path)}")
            except Exception: pass
            self._tab2_boot_loaded = False
            self._refresh_calc_order(); self._ensure_default_calc_loaded(); add_to_recents(path)
            messagebox.showinfo("Project", "Project opened."); self.nb.select(self.tab2)
        except Exception as e:
            messagebox.showerror("Open Project", f"{type(e).__name__}: {e}")

    def on_create_project(self):
        try:
            kv = self._collect_project_info(); proj_no = (kv.get("Project Number") or "").strip()
            if not proj_no:
                messagebox.showwarning("Create Project", "Enter a Project Number on Tab 1 first."); return
            path = project_filepath(proj_no); self.store = ProjectStore(path); self.store.open_or_create(); self.store.write_project_info(kv)
            self._fill_tab1_from_kv(kv); add_to_recents(path)
            self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv); self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
            try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(path)}")
            except Exception: pass
            self._tab2_boot_loaded = False; self._refresh_calc_order(); self._ensure_default_calc_loaded()
            messagebox.showinfo("Create Project", f"Created: {os.path.basename(path)}"); self.nb.select(self.tab2)
        except Exception as e:
            messagebox.showerror("Create Project", f"{type(e).__name__}: {e}")

    def on_open_recent(self):
        try:
            initial = PROJECT_DIR if os.path.isdir(PROJECT_DIR) else os.path.expanduser("~")
            path = filedialog.askopenfilename(title="Open Project Workbook", initialdir=initial, filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
            if not path:
                recents = get_recents(); path = recents[0] if recents else ""
                if not path: return
            self._load_project(path)
        except Exception as e:
            messagebox.showerror("Open Project", f"{type(e).__name__}: {e}")

    def on_save_project(self):
        try:
            kv = self._collect_project_info(); proj_no = (kv.get("Project Number") or "").strip()
            if not proj_no:
                messagebox.showwarning("Save Project", "Enter a Project Number on Tab 1 first."); return
            if not self.store:
                path = project_filepath(proj_no); self.store = ProjectStore(path); self.store.open_or_create()
            self.store.write_project_info(kv); self._fill_tab1_from_kv(kv); add_to_recents(self.store.path)
            self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv); self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
            try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(self.store.path)}")
            except Exception: pass
            messagebox.showinfo("Save Project", "Project information saved.")
            self.dirty_project = False
        except Exception as e:
            messagebox.showerror("Save Project", f"{type(e).__name__}: {e}")

    def on_search_project(self):
        try:
            q = (self.t1_search.get() or "").strip().lower()
            if not q:
                messagebox.showinfo("Search Project", "Type part of a filename into the search box first."); return
            pattern = os.path.join(PROJECT_DIR, "*.xlsx"); candidates = []
            try:
                for p in glob.glob(pattern):
                    if q in os.path.basename(p).lower(): candidates.append(p)
            except Exception: pass
            if not candidates:
                messagebox.showinfo("Search Project", "No matching .xlsx found in the projects folder."); return
            self._load_project(candidates[0])
        except Exception as e:
            messagebox.showerror("Search Project", f"{type(e).__name__}: {e}")

    # ---------- Close handler ----------
    def on_close(self):
        try:
            if getattr(self, "_autosave_job", None):
                self.after_cancel(self._autosave_job); self._autosave_job = None
        except Exception: pass
        try:
            if self.dirty_calc and self.store:
                if messagebox.askyesno("Unsaved changes", "Save current calculation before closing?"):
                    self.on_save_calc()
            if self.dirty_project and self.store:
                if messagebox.askyesno("Project changes", "Save project info before closing?"):
                    self.on_save_project()
        except Exception: pass
        self.destroy()
    # ---------- Tab 2 (Load & System) ----------
    def _build_tab2(self):
        f = self.tab2
        pad = {"padx": 10, "pady": 4}

        ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=12, sticky="w", **pad)

        self.t2_title = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad); self.t2_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
        self.t2_number = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Number", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad); self.t2_number.grid(row=1, column=4, sticky="we", **pad)
        self.t2_location = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad); self.t2_location.grid(row=1, column=6, sticky="we", **pad)
        self.t2_docno = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad); self.t2_docno.grid(row=2, column=1, sticky="we", **pad)
        self.t2_rev = ttk.Entry(f, width=10, style="App.TEntry", state="readonly"); ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad); self.t2_rev.grid(row=2, column=4, sticky="we", **pad)

        ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=12, sticky="we", padx=8, pady=6)
        ttk.Label(f, text="Calculation Input Data", style="App.Section.TLabel").grid(row=4, column=0, columnspan=12, sticky="w", **pad)

        ttk.Label(f, text="Calculation Number", style="App.TLabel").grid(row=5, column=0, sticky="w", **pad)
        self.t2_calcno = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_calcno.grid(row=5, column=1, sticky="w", **pad); self.t2_calcno.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Label(f, text="From Description", style="App.TLabel").grid(row=6, column=0, sticky="w", **pad); self.t2_from_desc = ttk.Entry(f, width=40, style="App.TEntry"); self.t2_from_desc.grid(row=6, column=1, columnspan=2, sticky="we", **pad); self.t2_from_desc.bind("<KeyRelease>", lambda e: self._schedule_autosave())
        ttk.Label(f, text="From Tag No", style="App.TLabel").grid(row=6, column=3, sticky="w", **pad); self.t2_from_tag = ttk.Entry(f, width=16, style="App.TEntry"); self.t2_from_tag.grid(row=6, column=4, sticky="we", **pad); self.t2_from_tag.bind("<KeyRelease>", lambda e: self._schedule_autosave())
        ttk.Label(f, text="Circuit ID No", style="App.TLabel").grid(row=6, column=5, sticky="w", **pad); self.t2_circuit = ttk.Entry(f, width=16, style="App.TEntry"); self.t2_circuit.grid(row=6, column=6, sticky="we", **pad); self.t2_circuit.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Label(f, text="To Description", style="App.TLabel").grid(row=7, column=0, sticky="w", **pad); self.t2_to_desc = ttk.Entry(f, width=40, style="App.TEntry"); self.t2_to_desc.grid(row=7, column=1, columnspan=2, sticky="we", **pad); self.t2_to_desc.bind("<KeyRelease>", lambda e: self._schedule_autosave())
        ttk.Label(f, text="To Tag No", style="App.TLabel").grid(row=7, column=3, sticky="w", **pad); self.t2_to_tag = ttk.Entry(f, width=16, style="App.TEntry"); self.t2_to_tag.grid(row=7, column=4, sticky="we", **pad); self.t2_to_tag.bind("<KeyRelease>", lambda e: self._schedule_autosave())
        ttk.Label(f, text="Cable Tag No", style="App.TLabel").grid(row=7, column=5, sticky="w", **pad); self.t2_cable = ttk.Entry(f, width=16, style="App.TEntry"); self.t2_cable.grid(row=7, column=6, sticky="we", **pad); self.t2_cable.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Separator(f, orient="horizontal").grid(row=8, column=0, columnspan=12, sticky="we", padx=8, pady=6)

        ttk.Label(f, text="System Data", style="App.Section.TLabel").grid(row=9, column=0, columnspan=12, sticky="w", **pad)
        ttk.Label(f, text="Phase (3ph or 1ph)", style="App.TLabel").grid(row=10, column=0, sticky="w", **pad)
        self.t2_phase = ttk.Combobox(f, values=["3","1"], width=6, state="readonly"); self.t2_phase.set("3"); self.t2_phase.grid(row=10, column=1, sticky="w", **pad); self.t2_phase.bind("<<ComboboxSelected>>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="System Voltage (L-L) (V)", style="App.TLabel").grid(row=11, column=0, sticky="w", **pad)
        self.t2_vll = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_vll.insert(0,"400"); self.t2_vll.grid(row=11, column=1, sticky="w", **pad); self.t2_vll.bind("<KeyRelease>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="Equipment Rating (kW)", style="App.TLabel").grid(row=12, column=0, sticky="w", **pad)
        self.t2_kw = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_kw.insert(0,"0.0"); self.t2_kw.grid(row=12, column=1, sticky="w", **pad); self.t2_kw.bind("<KeyRelease>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="Power Factor (-)", style="App.TLabel").grid(row=13, column=0, sticky="w", **pad)
        self.t2_pf = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_pf.insert(0,"0.9"); self.t2_pf.grid(row=13, column=1, sticky="w", **pad); self.t2_pf.bind("<KeyRelease>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="Efficiency (%)", style="App.TLabel").grid(row=14, column=0, sticky="w", **pad)
        self.t2_eff = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_eff.insert(0,"95.0"); self.t2_eff.grid(row=14, column=1, sticky="w", **pad); self.t2_eff.bind("<KeyRelease>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="Starting Current Factor (x)", style="App.TLabel").grid(row=15, column=0, sticky="w", **pad)
        self.t2_sf = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_sf.insert(0,"10.0"); self.t2_sf.grid(row=15, column=1, sticky="w", **pad); self.t2_sf.bind("<KeyRelease>", lambda e: (self._recalc(), self._schedule_autosave()))

        ttk.Label(f, text="Starting Power Factor (-)", style="App.TLabel").grid(row=16, column=0, sticky="w", **pad)
        self.t2_spf = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_spf.insert(0,"0.35"); self.t2_spf.grid(row=16, column=1, sticky="w", **pad); self.t2_spf.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Label(f, text="Circuit Length (m)", style="App.TLabel").grid(row=17, column=0, sticky="w", **pad)
        self.t2_len = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_len.insert(0,"50.0"); self.t2_len.grid(row=17, column=1, sticky="w", **pad); self.t2_len.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Label(f, text="Phase Voltage (V)", style="App.TLabel").grid(row=11, column=3, sticky="w", **pad)
        self.t2_vph = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t2_vph.grid(row=11, column=4, sticky="w", **pad)
        ttk.Label(f, text="Full Load Current", style="App.TLabel").grid(row=14, column=3, sticky="w", **pad)
        self.t2_flc = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t2_flc.grid(row=14, column=4, sticky="w", **pad)
        ttk.Label(f, text="Apparent Power", style="App.TLabel").grid(row=15, column=3, sticky="w", **pad)
        self.t2_kva = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t2_kva.grid(row=15, column=4, sticky="w", **pad)
        ttk.Label(f, text="Starting Current", style="App.TLabel").grid(row=16, column=3, sticky="w", **pad)
        self.t2_istart = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t2_istart.grid(row=16, column=4, sticky="w", **pad)

        ttk.Separator(f, orient="horizontal").grid(row=18, column=0, columnspan=12, sticky="we", padx=8, pady=6)
        ttk.Label(f, text="Protection Data", style="App.Section.TLabel").grid(row=19, column=0, columnspan=12, sticky="w", **pad)
        ttk.Label(f, text="Protective Device (A)", style="App.TLabel").grid(row=20, column=0, sticky="w", **pad)
        self.t2_pd = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_pd.insert(0,"0.0"); self.t2_pd.grid(row=20, column=1, sticky="w", **pad); self.t2_pd.bind("<KeyRelease>", lambda e: self._schedule_autosave())
        ttk.Label(f, text="Overload Setting (%)", style="App.TLabel").grid(row=20, column=3, sticky="w", **pad)
        self.t2_ol = ttk.Entry(f, width=12, style="App.TEntry"); self.t2_ol.insert(0,"100.0"); self.t2_ol.grid(row=20, column=4, sticky="w", **pad); self.t2_ol.bind("<KeyRelease>", lambda e: self._schedule_autosave())

        ttk.Separator(f, orient="horizontal").grid(row=21, column=0, columnspan=12, sticky="we", **pad)
        ttk.Button(f, text="New Calculation", command=self.on_new_calc, style="App.TButton").grid(row=22, column=0, sticky="w", **pad)
        ttk.Button(f, text="Save/Update Calculation", command=self.on_save_calc, style="App.TButton").grid(row=22, column=1, sticky="w", **pad)
        self.t2_search = ttk.Entry(f, width=36, style="App.TEntry"); self.t2_search.grid(row=22, column=3, sticky="we", **pad)
        ttk.Button(f, text="Search Calculation", command=self.on_search_calc, style="App.TButton").grid(row=22, column=4, sticky="w", **pad)
        self.btn_calc_prev = ttk.Button(f, text="Prev", command=self.on_prev_calc, style="App.TButton"); self.btn_calc_prev.grid(row=22, column=6, sticky="e", **pad)
        self.btn_calc_next = ttk.Button(f, text="Next", command=self.on_next_calc, style="App.TButton"); self.btn_calc_next.grid(row=22, column=7, sticky="e", **pad)
        for c in (1,2,4,6): f.grid_columnconfigure(c, weight=1)

        for w in [self.t2_phase, self.t2_vll, self.t2_kw, self.t2_pf, self.t2_eff, self.t2_sf,
                  self.t2_spf, self.t2_pd, self.t2_ol, self.t2_len, self.t2_calcno,
                  self.t2_from_desc, self.t2_to_desc, self.t2_from_tag, self.t2_to_tag,
                  self.t2_circuit, self.t2_cable]:
            self._install_text_bindings(w)

    def _recalc(self):
        try:
            phase = self.t2_phase.get().strip() or "3"
            vll = float(self.t2_vll.get()); kw = float(self.t2_kw.get())
            pf = float(self.t2_pf.get()); eff = float(self.t2_eff.get()); sf = float(self.t2_sf.get())
        except Exception:
            for e in [self.t2_vph, self.t2_kva, self.t2_flc, self.t2_istart]:
                e.configure(state="normal"); e.delete(0,"end"); e.configure(state="readonly")
            return
        vph = calc_phase_voltage(vll, phase)
        kva = calc_kva_kw_pf_eff(kw, pf, eff)
        flc = calc_flc(kw, vll, phase, pf, eff)
        ist = calc_start_current(flc, sf)
        for (e, val) in [(self.t2_vph, f"{vph:.2f}"), (self.t2_kva, f"{kva:.3f}"), (self.t2_flc, f"{flc:.2f}"), (self.t2_istart, f"{ist:.2f}")]:
            e.configure(state="normal"); e.delete(0,"end"); e.insert(0, val); e.configure(state="readonly")
        self.dirty_calc = True

    # --- calc list/navigation ---
    def _refresh_calc_order(self):
        try:
            self.calc_order = self.store.list_calc_numbers() if self.store else []
        except Exception:
            self.calc_order = []
        current_no = ""
        try:
            current_no = (self.t2_calcno.get().strip() if getattr(self, "t2_calcno", None) else "")
        except Exception: pass
        if self.calc_order:
            self.calc_index = self.calc_order.index(current_no) if (current_no and current_no in self.calc_order) else 0
        else:
            self.calc_index = -1
        self._update_calc_nav_buttons()

    def _update_calc_nav_buttons(self):
        total = len(getattr(self, "calc_order", []) or [])
        idx = getattr(self, "calc_index", -1)
        for btn in [getattr(self, "btn_calc_prev", None), getattr(self, "btn3_prev", None), getattr(self, "btn4_prev", None), getattr(self, "btn5_prev", None)]:
            if btn: btn.state(["!disabled"] if idx > 0 else ["disabled"])
        for btn in [getattr(self, "btn_calc_next", None), getattr(self, "btn3_next", None), getattr(self, "btn4_next", None), getattr(self, "btn5_next", None)]:
            if btn: btn.state(["!disabled"] if (total and idx < total - 1) else ["disabled"])

    def _ensure_default_calc_loaded(self):
        if not self.store:
            self._tab2_boot_loaded = True; return
        self._refresh_calc_order()
        if not self.calc_order:
            try:
                self.store.upsert_calc_row({"CalculationNumber": "0001"})
                self._refresh_calc_order()
            except Exception: pass
        hit = None
        try: hit = self.store.read_calc_by_number("0001")
        except Exception: hit = None
        if hit:
            idx, row = hit
            self.calc_index = idx; self._fill_from_row(row)
        else:
            row = self.store.read_calc_by_index(0)
            if row: self.calc_index = 0; self._fill_from_row(row)
        self._tab2_boot_loaded = True; self._update_calc_nav_buttons()

    def _on_tab_changed(self, _event=None):
        try:
            self._update_calc_nav_buttons()
            current_index = self.nb.index("current")
            if current_index in (1,2,3,4):
                if not self._tab2_boot_loaded:
                    self._ensure_default_calc_loaded()
                calcno = (self.t2_calcno.get().strip() if getattr(self, "t2_calcno", None) else "")
                if calcno:
                    for w in [getattr(self, "t3_calcno", None), getattr(self, "t4_calcno", None), getattr(self, "t5_calcno", None), getattr(self, "t6_calcno", None)]:
                        if w is not None:
                            w.configure(state="normal"); w.delete(0, "end"); w.insert(0, calcno); w.configure(state="readonly")
        except Exception:
            pass

    def _load_calc_by_index(self, idx: int):
        if not self.store: return
        row = self.store.read_calc_by_index(idx)
        if not row: return
        self.calc_index = idx
        self._fill_from_row(row)
        self._update_calc_nav_buttons()

    def _next_calc_number(self) -> str:
        try:
            self._refresh_calc_order()
            nums = [int(n) for n in self.calc_order if str(n).isdigit()]
            nxt = (max(nums) + 1) if nums else 1
        except Exception:
            nxt = 1
        return f"{nxt:04d}"

    def on_new_calc(self):
        if not self.store:
            messagebox.showwarning("New Calculation", "Save the project first (Tab 1)."); return
        new_no = self._next_calc_number()
        for e in [self.t2_from_desc, self.t2_from_tag, self.t2_circuit, self.t2_to_desc, self.t2_to_tag, self.t2_cable, self.t2_kw, self.t2_pf, self.t2_eff, self.t2_sf, self.t2_spf, self.t2_pd, self.t2_ol, self.t2_len]:
            try: e.delete(0, "end")
            except Exception: pass
        self.t2_calcno.delete(0, "end"); self.t2_calcno.insert(0, new_no)
        for e in [self.t2_vph, self.t2_kva, self.t2_flc, self.t2_istart]:
            e.configure(state="normal"); e.delete(0, "end"); e.configure(state="readonly")
        for w in [self.t3_calcno, self.t4_calcno, self.t5_calcno, self.t6_calcno]:
            w.configure(state="normal"); w.delete(0, "end"); w.insert(0, new_no); w.configure(state="readonly")
        self._refresh_calc_order(); self.calc_index = len(self.calc_order); self._update_calc_nav_buttons()

    def on_search_calc(self):
        if not self.store:
            messagebox.showwarning("Search Calculation", "Open or save a project first."); return
        q = (self.t2_search.get() or "").strip()
        if not q:
            messagebox.showinfo("Search", "Enter text to search in calculations."); return
        hits = self.store.search_calculations(q)
        if not hits:
            messagebox.showinfo("Search", "No matching calculation found."); return
        idx, row = hits[0]
        self.calc_index = idx; self._fill_from_row(row); self._update_calc_nav_buttons()

    def on_prev_calc(self):
        if not self.store: return
        self._refresh_calc_order()
        if self.calc_index > 0:
            self._load_calc_by_index(self.calc_index - 1)

    def on_next_calc(self):
        if not self.store: return
        self._refresh_calc_order()
        if self.calc_order and self.calc_index < len(self.calc_order) - 1:
            self._load_calc_by_index(self.calc_index + 1)
        else:
            self._update_calc_nav_buttons()

    # --- collect/save & fill ---
    def _collect_calc_dict(self) -> Dict[str, object]:
        calcno = (self.t2_calcno.get().strip() if getattr(self, "t2_calcno", None) else "")
        phase = (self.t2_phase.get().strip() if getattr(self, "t2_phase", None) else "3")
        vll   = self._to_float(self.t2_vll.get() if getattr(self, "t2_vll", None) else 0.0)
        kw    = self._to_float(self.t2_kw.get() if getattr(self, "t2_kw", None) else 0.0)
        pf    = self._to_float(self.t2_pf.get() if getattr(self, "t2_pf", None) else 0.0)
        eff   = self._to_float(self.t2_eff.get() if getattr(self, "t2_eff", None) else 0.0)
        sf    = self._to_float(self.t2_sf.get() if getattr(self, "t2_sf", None) else 0.0)

        vph  = calc_phase_voltage(vll, phase) if vll else 0.0
        kva  = calc_kva_kw_pf_eff(kw, pf, eff) if kw and pf and eff else 0.0
        flc  = calc_flc(kw, vll, phase, pf, eff) if kw and vll and pf and eff else 0.0
        ist  = calc_start_current(flc, sf) if flc and sf else 0.0

        payload = {
            "CalculationNumber": calcno,
            "FromDescription": self.t2_from_desc.get().strip(),
            "ToDescription":   self.t2_to_desc.get().strip(),
            "FromTag":         self.t2_from_tag.get().strip(),
            "ToTag":           self.t2_to_tag.get().strip(),
            "CircuitID":       self.t2_circuit.get().strip(),
            "CableTag":        self.t2_cable.get().strip(),
            "Phase": phase, "SystemVoltageLL_V": vll, "PhaseVoltage_V": vph,
            "EquipmentRating_kW": kw, "PowerFactor": pf, "Efficiency_pct": eff,
            "StartFactor": sf, "StartPF": self._to_float(self.t2_spf.get()), "ApparentPower_kVA": kva,
            "FullLoadCurrent_A": flc, "StartingCurrent_A": ist,
            "ProtectiveDevice_A": self._to_float(self.t2_pd.get()), "OverloadSetting_pct": self._to_float(self.t2_ol.get()),
            "CircuitLength_m": self._to_float(self.t2_len.get()),
        }
        # Tab 3/4/5 values are added when those tabs run their own save/update
        # but we'll include whatever is already present (safe if empty):
        for name in [
            "TypeOfCable","TypeOfInstallation","CoreType","Formation",
            "DepthOfCable","DepthCF","SoilThermalResistivity","SoilResistivityCF",
            "NumberOfCircuits","GroundOrDuctsCF","SpacingBetweenCircuits",
            "CableManagement","NumberOfTraysLadders","NumberOfCablesPerTray",
            "InAirDeratingFactor","AmbientTemperature_C","TemperatureDeratingFactor","OverallDeratingFactor",
            "ProtectionType","I2_Factor","CableRatedCurrent_Iz_single","CablesInParallel_n","TotalCableRating_Iz","OverloadCheck_OK",
            "Cable_r_ohm_per_km","Cable_x_ohm_per_km","SteadyVD_V","SteadyVD_pct","SteadyVD_limit_pct","SteadyVD_OK",
            "StartingVD_V","StartingVD_pct","StartingVD_limit_pct","StartingVD_OK",
        ]:
            try:
                payload[name] = payload.get(name, getattr(self, f"_payload_{name}", ""))
            except Exception:
                pass
        return payload

    def on_save_calc(self):
        if not self.store:
            messagebox.showwarning("Save Calculation", "Save the project first (Tab 1) to create the workbook."); return
        try:
            payload = self._collect_calc_dict()
            if not payload.get("CalculationNumber"):
                messagebox.showwarning("Calculation Number", "Please enter a Calculation Number (e.g., 0001)."); return
            self.store.upsert_calc_row(payload)
            self._refresh_calc_order()
            if payload["CalculationNumber"] in self.calc_order:
                self.calc_index = self.calc_order.index(payload["CalculationNumber"])
            self._update_calc_nav_buttons()
            messagebox.showinfo("Saved", f"Calculation {payload['CalculationNumber']} saved/updated.")
            self.dirty_calc = False
        except Exception as e:
            messagebox.showerror("Save Calculation", f"{type(e).__name__}: {e}")

    def _fill_from_row(self, row: Dict[str, object]):
        def _txt(entry, val):
            if entry is None: return
            ro = (entry.cget("state") == "readonly")
            if ro: entry.configure(state="normal")
            entry.delete(0, "end"); entry.insert(0, "" if val is None else str(val))
            if ro: entry.configure(state="readonly")
        def _combo(entry, val):
            if entry is None: return
            try: entry.configure(state="readonly"); entry.set("" if val is None else str(val))
            except Exception:
                try: entry.set("" if val is None else str(val))
                except Exception: pass

        for key, widget_name in [
            ("CalculationNumber","t2_calcno"), ("FromDescription","t2_from_desc"),
            ("FromTag","t2_from_tag"), ("CircuitID","t2_circuit"),
            ("ToDescription","t2_to_desc"), ("ToTag","t2_to_tag"), ("CableTag","t2_cable"),
            ("SystemVoltageLL_V","t2_vll"), ("EquipmentRating_kW","t2_kw"),
            ("PowerFactor","t2_pf"), ("Efficiency_pct","t2_eff"),
            ("StartFactor","t2_sf"), ("StartPF","t2_spf"),
            ("ProtectiveDevice_A","t2_pd"), ("OverloadSetting_pct","t2_ol"),
            ("CircuitLength_m","t2_len")
        ]:
            _txt(getattr(self, widget_name, None), row.get(key))
        _combo(getattr(self, "t2_phase", None), row.get("Phase"))
        for key, widget_name in [("PhaseVoltage_V","t2_vph"), ("ApparentPower_kVA","t2_kva"), ("FullLoadCurrent_A","t2_flc"), ("StartingCurrent_A","t2_istart")]:
            _txt(getattr(self, widget_name, None), row.get(key))
        try: self._recalc()
        except Exception: pass
# ---------- Tab 3 (Installation & Env) ----------
def _build_tab3(self):
    f = self.tab3
    pad = {"padx": 10, "pady": 4}

    ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=12, sticky="w", **pad)
    self.t3_title = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad); self.t3_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
    self.t3_number = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Number", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad); self.t3_number.grid(row=1, column=4, sticky="we", **pad)
    self.t3_location = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad); self.t3_location.grid(row=1, column=6, sticky="we", **pad)
    self.t3_docno = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad); self.t3_docno.grid(row=2, column=1, sticky="we", **pad)
    self.t3_rev = ttk.Entry(f, width=10, style="App.TEntry", state="readonly"); ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad); self.t3_rev.grid(row=2, column=4, sticky="we", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=12, sticky="we", padx=8, pady=6)

    ttk.Label(f, text="Cable Input Data", style="App.Section.TLabel").grid(row=4, column=0, columnspan=12, sticky="w", **pad)
    ttk.Label(f, text="Calculation Number", style="App.TLabel").grid(row=5, column=0, sticky="w", **pad)
    self.t3_calcno = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); self.t3_calcno.grid(row=5, column=1, sticky="w", **pad)

    ttk.Label(f, text="From Description", style="App.TLabel").grid(row=6, column=0, sticky="w", **pad)
    self.t3_from_desc = ttk.Entry(f, width=40, style="App.TEntry", state="readonly"); self.t3_from_desc.grid(row=6, column=1, columnspan=2, sticky="we", **pad)
    ttk.Label(f, text="From Tag No", style="App.TLabel").grid(row=6, column=3, sticky="w", **pad)
    self.t3_from_tag = ttk.Entry(f, width=16, style="App.TEntry", state="readonly"); self.t3_from_tag.grid(row=6, column=4, sticky="we", **pad)
    ttk.Label(f, text="Circuit ID No", style="App.TLabel").grid(row=6, column=5, sticky="w", **pad)
    self.t3_circuit = ttk.Entry(f, width=16, style="App.TEntry", state="readonly"); self.t3_circuit.grid(row=6, column=6, sticky="we", **pad)

    ttk.Label(f, text="To Description", style="App.TLabel").grid(row=7, column=0, sticky="w", **pad)
    self.t3_to_desc = ttk.Entry(f, width=40, style="App.TEntry", state="readonly"); self.t3_to_desc.grid(row=7, column=1, columnspan=2, sticky="we", **pad)
    ttk.Label(f, text="To Tag No", style="App.TLabel").grid(row=7, column=3, sticky="w", **pad)
    self.t3_to_tag = ttk.Entry(f, width=16, style="App.TEntry", state="readonly"); self.t3_to_tag.grid(row=7, column=4, sticky="we", **pad)
    ttk.Label(f, text="Cable Tag No", style="App.TLabel").grid(row=7, column=5, sticky="w", **pad)
    self.t3_cable = ttk.Entry(f, width=16, style="App.TEntry", state="readonly"); self.t3_cable.grid(row=7, column=6, sticky="we", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=8, column=0, columnspan=12, sticky="we", **pad)

    ttk.Label(f, text="Cable Information", style="App.Section.TLabel").grid(row=9, column=0, columnspan=12, sticky="w", **pad)

    ttk.Label(f, text="Type of Cable", style="App.TLabel").grid(row=10, column=0, sticky="w", **pad)
    self.t3_type_cable = ttk.Combobox(f, values=["PVC", "XLPE", "MICC", "EPR"], width=22, state="readonly"); self.t3_type_cable.grid(row=10, column=1, sticky="w", **pad)
    self.t3_type_cable.bind("<<ComboboxSelected>>", lambda e: (self._temp_lookup_and_set(), self._schedule_autosave()))

    ttk.Label(f, text="Type of Installation", style="App.TLabel").grid(row=10, column=3, sticky="w", **pad)
    self.t3_type_install = ttk.Combobox(f, values=["Direct in Ground", "In Ducts", "In Air"], width=22, state="readonly"); self.t3_type_install.grid(row=10, column=4, sticky="w", **pad)
    self.t3_type_install.bind("<<ComboboxSelected>>", lambda e: (self._update_air_ui(), self._schedule_autosave()))

    ttk.Label(f, text="Core Type", style="App.TLabel").grid(row=10, column=5, sticky="w", **pad)
    self.t3_core_type = ttk.Combobox(f, values=["Multicore", "Single Core"], width=22, state="readonly"); self.t3_core_type.set("Multicore"); self.t3_core_type.grid(row=10, column=6, sticky="w", **pad)
    self.t3_core_type.bind("<<ComboboxSelected>>", lambda e: (self._on_core_type_change(), self._update_air_ui(), self._schedule_autosave()))

    ttk.Label(f, text="Formation", style="App.TLabel").grid(row=10, column=7, sticky="w", **pad)
    self.t3_formation = ttk.Combobox(f, values=["Flat", "Trefoil"], width=22, state="disabled"); self.t3_formation.grid(row=10, column=8, sticky="w", **pad)
    self.t3_formation.bind("<<ComboboxSelected>>", lambda e: (self._update_air_ui(), self._schedule_autosave()))

    ttk.Separator(f, orient="horizontal").grid(row=11, column=0, columnspan=12, sticky="we", padx=8, pady=6)

    ttk.Label(f, text="Cable Installed Direct in GROUND or IN DUCTS", style="App.Section.TLabel").grid(row=12, column=0, columnspan=12, sticky="w", **pad)
    ttk.Label(f, text="Depth of Cable", style="App.TLabel").grid(row=13, column=0, sticky="w", **pad); self.t3_depth = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_depth.grid(row=13, column=1, sticky="w", **pad); self.t3_depth.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Correction Factor for Depth of Cable", style="App.TLabel").grid(row=13, column=3, sticky="w", **pad); self.t3_depth_cf = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_depth_cf.grid(row=13, column=4, sticky="w", **pad); self.t3_depth_cf.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Soil Thermal Resistivity", style="App.TLabel").grid(row=14, column=0, sticky="w", **pad); self.t3_soil_r = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_soil_r.grid(row=14, column=1, sticky="w", **pad); self.t3_soil_r.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Soil Resistivity Correction Factor", style="App.TLabel").grid(row=14, column=3, sticky="w", **pad); self.t3_soil_cf = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_soil_cf.grid(row=14, column=4, sticky="w", **pad); self.t3_soil_cf.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Number of Circuits being installed", style="App.TLabel").grid(row=15, column=0, sticky="w", **pad); self.t3_circuits = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_circuits.grid(row=15, column=1, sticky="w", **pad); self.t3_circuits.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Correction factor in Grnd/Ducts", style="App.TLabel").grid(row=15, column=3, sticky="w", **pad); self.t3_ground_cf = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_ground_cf.grid(row=15, column=4, sticky="w", **pad); self.t3_ground_cf.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Spacing Between the Circuits", style="App.TLabel").grid(row=16, column=0, sticky="w", **pad); self.t3_spacing = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_spacing.grid(row=16, column=1, sticky="w", **pad); self.t3_spacing.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))

    ttk.Separator(f, orient="horizontal").grid(row=17, column=0, columnspan=12, sticky="we", **pad)

    ttk.Label(f, text="Cable Installed Direct in FREE AIR", style="App.Section.TLabel").grid(row=18, column=0, columnspan=12, sticky="w", **pad)
    ttk.Label(f, text="Ambient Temperature (°C)", style="App.TLabel").grid(row=19, column=0, sticky="w", **pad)
    self.t3_temp = ttk.Combobox(f, values=[str(k) for k in [25,30,35,40,45,50,55]], width=10, state="disabled"); self.t3_temp.grid(row=19, column=1, sticky="w", **pad); self.t3_temp.bind("<<ComboboxSelected>>", lambda e: (self._temp_lookup_and_set(), self._schedule_autosave()))
    ttk.Label(f, text="Temperature Derating Factor", style="App.TLabel").grid(row=19, column=3, sticky="w", **pad); self.t3_temp_df = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_temp_df.grid(row=19, column=4, sticky="w", **pad); self.t3_temp_df.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="Cable Management", style="App.TLabel").grid(row=20, column=0, sticky="w", **pad); self.t3_cable_mgmt = ttk.Combobox(f, values=[], width=25, state="disabled"); self.t3_cable_mgmt.grid(row=20, column=1, columnspan=3, sticky="we", **pad); self.t3_cable_mgmt.bind("<<ComboboxSelected>>", lambda e: (self._populate_tray_cable_options(), self._schedule_autosave()))
    ttk.Label(f, text="No of Trays /Ladders", style="App.TLabel").grid(row=21, column=0, sticky="w", **pad); self.t3_trays = ttk.Combobox(f, values=[], width=18, state="disabled"); self.t3_trays.grid(row=21, column=1, sticky="w", **pad); self.t3_trays.bind("<<ComboboxSelected>>", lambda e: (self._air_lookup_and_set(), self._schedule_autosave()))
    ttk.Label(f, text="In Air Cable Derating Factor", style="App.TLabel").grid(row=21, column=3, sticky="w", **pad); self.t3_air_df = ttk.Entry(f, width=18, style="App.TEntry"); self.t3_air_df.grid(row=21, column=4, sticky="w", **pad); self.t3_air_df.bind("<KeyRelease>", lambda e: (self._recalc_odf(), self._schedule_autosave()))
    ttk.Label(f, text="No of Cables per Tray/Ladder", style="App.TLabel").grid(row=22, column=0, sticky="w", **pad); self.t3_cables_per_tray = ttk.Combobox(f, values=[], width=18, state="disabled"); self.t3_cables_per_tray.grid(row=22, column=1, sticky="w", **pad); self.t3_cables_per_tray.bind("<<ComboboxSelected>>", lambda e: (self._air_lookup_and_set(), self._schedule_autosave()))

    ttk.Separator(f, orient="horizontal").grid(row=23, column=0, columnspan=12, sticky="we", padx=8, pady=8)
    ttk.Label(f, text="Overall Derating Factor (ODF)", style="App.Section.TLabel").grid(row=24, column=0, columnspan=2, sticky="w", **pad)
    self.t3_odf = ttk.Entry(f, width=18, style="App.TEntry", state="readonly"); self.t3_odf.grid(row=24, column=1, sticky="w", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=25, column=0, columnspan=12, sticky="we", **pad)
    self.btn3_save = ttk.Button(f, text="Save/Update Calculation", style="App.TButton", command=self.on_save_calc); self.btn3_save.grid(row=26, column=0, sticky="w", **pad)
    self.btn3_prev = ttk.Button(f, text="Prev", style="App.TButton", command=self.on_prev_calc); self.btn3_prev.grid(row=26, column=6, sticky="e", **pad)
    self.btn3_next = ttk.Button(f, text="Next", style="App.TButton", command=self.on_next_calc); self.btn3_next.grid(row=26, column=7, sticky="e", **pad)

    for c in (1,2,4,6,8): f.grid_columnconfigure(c, weight=1)
    self.t3_core_type.bind("<<ComboboxSelected>>", lambda e: self._on_core_type_change())
    self._on_core_type_change(); self._recalc_odf(); self._update_air_ui()

def _on_core_type_change(self):
    ct = (self.t3_core_type.get() or "").strip()
    if ct == "Single Core":
        self.t3_formation.configure(state="readonly")
        if not self.t3_formation.get(): self.t3_formation.set("Flat")
    else:
        self.t3_formation.set(""); self.t3_formation.configure(state="disabled")
    try: self._update_air_ui()
    except Exception: pass

def _truncate_label(self, s: str, maxlen: int = 50) -> str:
    s = str(s or ""); return s if len(s) <= maxlen else (s[:maxlen-1] + "…")

def _set_cable_mgmt_options(self, full_options: List[str]):
    self.cm_label_map = {}; display_opts = []
    for key in full_options:
        lab = self._truncate_label(key, 50)
        while lab in self.cm_label_map and self.cm_label_map[lab] != key:
            lab = self._truncate_label(lab + " *", 50)
        self.cm_label_map[lab] = key; display_opts.append(lab)
    self.t3_cable_mgmt.configure(values=display_opts)

def _resolve_cable_mgmt_key(self, display: str) -> str:
    return self.cm_label_map.get(display, display)

def _update_air_ui(self):
    install = (self.t3_type_install.get() or "").strip()
    core    = (self.t3_core_type.get() or "").strip()
    is_air = (install == "In Air")
    for w in [self.t3_cable_mgmt, self.t3_trays, self.t3_cables_per_tray, self.t3_temp]:
        w.configure(state=("readonly" if is_air else "disabled"))
    if not is_air:
        for w in [self.t3_cable_mgmt, self.t3_trays, self.t3_cables_per_tray, self.t3_temp]:
            try: w.set("")
            except Exception: pass
        for e in [self.t3_air_df, self.t3_temp_df]:
            e.delete(0,"end"); e.insert(0,"1.00")
        self._recalc_odf(); return
    opts = SC_CABLE_MGMT_OPTIONS if core == "Single Core" else list(FREE_AIR_MC.keys())
    self._set_cable_mgmt_options(opts)
    if not self.t3_cable_mgmt.get() and self.t3_cable_mgmt.cget("values"):
        self.t3_cable_mgmt.set(self.t3_cable_mgmt.cget("values")[0])
    if not self.t3_temp.get():
        try: self.t3_temp.set("30")
        except Exception: pass
    self._temp_lookup_and_set(); self._populate_tray_cable_options()

def _current_sc_table(self) -> Dict[int, Dict[int, float]]:
    mgmt_display = (self.t3_cable_mgmt.get() or "").strip()
    mgmt = self._resolve_cable_mgmt_key(mgmt_display)
    formation = (self.t3_formation.get() or "Flat").strip()
    if formation == "Trefoil":
        return FREE_AIR_SC_TREFOIL.get(mgmt, {})
    table = FREE_AIR_SC_FLAT.get(mgmt)
    if table is None and mgmt == "Horizontal Ladder System (Spaced)":
        return FREE_AIR_SC_TREFOIL.get(mgmt, {})
    return table or {}

def _populate_tray_cable_options(self):
    name_display = (self.t3_cable_mgmt.get() or "").strip()
    name = self._resolve_cable_mgmt_key(name_display)
    core = (self.t3_core_type.get() or "").strip()
    table = self._current_sc_table() if core == "Single Core" else FREE_AIR_MC.get(name, {})
    trays_list, cables_list = [], []
    if table:
        trays_list = sorted(table.keys())
        cable_set = set(); [cable_set.update(d.keys()) for d in table.values()]
        cables_list = sorted(cable_set)
    self.t3_trays.configure(values=trays_list)
    self.t3_cables_per_tray.configure(values=cables_list)
    if trays_list and not self.t3_trays.get(): self.t3_trays.set(str(trays_list[0]))
    if cables_list and not self.t3_cables_per_tray.get(): self.t3_cables_per_tray.set(str(cables_list[0]))
    self._air_lookup_and_set()

def _air_lookup_and_set(self):
    try:
        core = (self.t3_core_type.get() or "").strip()
        name = self._resolve_cable_mgmt_key((self.t3_cable_mgmt.get() or "").strip())
        trays = int(self.t3_trays.get()); cables = int(self.t3_cables_per_tray.get())
        table = self._current_sc_table() if core == "Single Core" else FREE_AIR_MC.get(name, {})
        factor = float(table.get(trays, {}).get(cables, 1.0))
    except Exception:
        factor = 1.0
    self.t3_air_df.delete(0, "end"); self.t3_air_df.insert(0, f"{factor:.2f}"); self._recalc_odf()

def _temp_lookup_and_set(self):
    try: t = int((self.t3_temp.get() or "50").strip())
    except Exception: t = 50
    ctype = (self.t3_type_cable.get() or "").upper()
    row_key = "PVC" if "PVC" in ctype else ("XLPE" if "XLPE" in ctype else None)
    fac = 1.0
    try:
        if row_key is not None: fac = float(TEMP_AIR_FACTORS.get(row_key, {}).get(t, 1.0))
    except Exception: fac = 1.0
    self.t3_temp_df.delete(0, "end"); self.t3_temp_df.insert(0, f"{fac:.2f}"); self._recalc_odf()

def _recalc_odf(self):
    def fac(entry):
        txt = entry.get().strip();
        if not txt: return 1.0
        try: return float(txt)
        except Exception: return 1.0
    odf = fac(self.t3_depth_cf) * fac(self.t3_soil_cf) * fac(self.t3_ground_cf) * fac(self.t3_air_df) * fac(self.t3_temp_df)
    self.t3_odf.configure(state="normal"); self.t3_odf.delete(0,"end"); self.t3_odf.insert(0, f"{odf:.3f}"); self.t3_odf.configure(state="readonly")
# ---------- Tab 4 (Overload Check) ----------
def _build_tab4(self):
    f = self.tab4
    pad = {"padx":10, "pady":4}
    ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=12, sticky="w", **pad)
    self.t4_title = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad); self.t4_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
    self.t4_number = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Number", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad); self.t4_number.grid(row=1, column=4, sticky="we", **pad)
    self.t4_location = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad); self.t4_location.grid(row=1, column=6, sticky="we", **pad)
    self.t4_docno = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad); self.t4_docno.grid(row=2, column=1, sticky="we", **pad)
    self.t4_rev = ttk.Entry(f, width=10, style="App.TEntry", state="readonly"); ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad); self.t4_rev.grid(row=2, column=4, sticky="we", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=12, sticky="we", padx=8, pady=6)
    ttk.Label(f, text="Overload Protection Check (IEC 60364)", style="App.Section.TLabel").grid(row=4, column=0, columnspan=12, sticky="w", **pad)

    ttk.Label(f, text="Calculation Number", style="App.TLabel").grid(row=5, column=0, sticky="w", **pad); self.t4_calcno = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); self.t4_calcno.grid(row=5, column=1, sticky="w", **pad)
    ttk.Label(f, text="Protection Type", style="App.TLabel").grid(row=6, column=0, sticky="w", **pad)
    self.t4_ptype = ttk.Combobox(f, values=list(PROTECTION_I2_MULT.keys()), width=28, state="readonly"); self.t4_ptype.grid(row=6, column=1, sticky="w", **pad)
    self.t4_ptype.bind("<<ComboboxSelected>>", lambda e: (self._on_ptype_change(), self._recalc_overload(), self._schedule_autosave()))

    ttk.Label(f, text="I2 factor (×In)", style="App.TLabel").grid(row=6, column=3, sticky="w", **pad); self.t4_i2 = ttk.Entry(f, width=10, style="App.TEntry"); self.t4_i2.grid(row=6, column=4, sticky="w", **pad); self.t4_i2.bind("<KeyRelease>", lambda e: (self._recalc_overload(), self._schedule_autosave()))
    ttk.Label(f, text="Cable Iz (single) (A)", style="App.TLabel").grid(row=7, column=0, sticky="w", **pad); self.t4_iz_single = ttk.Entry(f, width=14, style="App.TEntry"); self.t4_iz_single.grid(row=7, column=1, sticky="w", **pad); self.t4_iz_single.bind("<KeyRelease>", lambda e: (self._recalc_overload(), self._schedule_autosave()))
    ttk.Label(f, text="Parallel Cables (n)", style="App.TLabel").grid(row=7, column=3, sticky="w", **pad); self.t4_n_parallel = ttk.Combobox(f, values=[str(i) for i in range(1,7)], width=10, state="readonly"); self.t4_n_parallel.set("1"); self.t4_n_parallel.grid(row=7, column=4, sticky="w", **pad); self.t4_n_parallel.bind("<<ComboboxSelected>>", lambda e: (self._recalc_overload(), self._schedule_autosave()))

    ttk.Label(f, text="Design Current Ib (A)", style="App.TLabel").grid(row=8, column=0, sticky="w", **pad); self.t4_ib = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t4_ib.grid(row=8, column=1, sticky="w", **pad)
    ttk.Label(f, text="Device Rating In (A)", style="App.TLabel").grid(row=8, column=3, sticky="w", **pad); self.t4_in = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t4_in.grid(row=8, column=4, sticky="w", **pad)

    ttk.Label(f, text="Total Cable Rating Iz (A)", style="App.TLabel").grid(row=9, column=0, sticky="w", **pad); self.t4_iz_total = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t4_iz_total.grid(row=9, column=1, sticky="w", **pad)
    ttk.Label(f, text="Overload Check Result", style="App.TLabel").grid(row=9, column=3, sticky="w", **pad); self.t4_ok = ttk.Entry(f, width=20, style="App.TEntry", state="readonly"); self.t4_ok.grid(row=9, column=4, sticky="w", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=10, column=0, columnspan=12, sticky="we", padx=8, pady=8)
    ttk.Button(f, text="Save/Update Calculation", style="App.TButton", command=self.on_save_calc).grid(row=11, column=0, sticky="w", **pad)
    self.btn4_prev = ttk.Button(f, text="Prev", style="App.TButton", command=self.on_prev_calc); self.btn4_prev.grid(row=11, column=6, sticky="e", **pad)
    self.btn4_next = ttk.Button(f, text="Next", style="App.TButton", command=self.on_next_calc); self.btn4_next.grid(row=11, column=7, sticky="e", **pad)

    for c in (1,4,6): f.grid_columnconfigure(c, weight=1)

def _on_ptype_change(self):
    ptype = (self.t4_ptype.get() or "").strip()
    fac = PROTECTION_I2_MULT.get(ptype)
    if fac is not None:
        self.t4_i2.delete(0,"end"); self.t4_i2.insert(0, f"{fac:.2f}")

def _recalc_overload(self):
    try:
        Ib = float(self.t2_flc.get()) if self.t2_flc.get() else 0.0
        In = float(self.t2_pd.get()) if self.t2_pd.get() else 0.0
        odf = float(self.t3_odf.get()) if self.t3_odf.get() else 1.0
        iz_single = float(self.t4_iz_single.get() or 0.0)
        npar = int(self.t4_n_parallel.get() or 1)
        i2fac = float(self.t4_i2.get() or 1.45)
    except Exception:
        Ib=In=odf=iz_single=0.0; npar=1; i2fac=1.45
    try:
        self.t4_ib.configure(state="normal"); self.t4_ib.delete(0,"end"); self.t4_ib.insert(0, f"{Ib:.2f}"); self.t4_ib.configure(state="readonly")
        self.t4_in.configure(state="normal"); self.t4_in.delete(0,"end"); self.t4_in.insert(0, f"{In:.0f}"); self.t4_in.configure(state="readonly")
    except Exception: pass
    Iz = iz_single * npar * odf
    I2 = i2fac * In
    try:
        self.t4_iz_total.configure(state="normal"); self.t4_iz_total.delete(0,"end"); self.t4_iz_total.insert(0, f"{Iz:.2f}"); self.t4_iz_total.configure(state="readonly")
    except Exception: pass
    cond1 = (Ib <= In <= Iz) if (In>0 and Iz>0) else False
    cond2 = (I2 <= 1.45 * Iz) if (I2>0 and Iz>0) else False
    res = "OK" if (cond1 and cond2) else ("FAIL: Ib≤In≤Iz" if not cond1 else "FAIL: I2≤1.45×Iz")
    try:
        self.t4_ok.configure(state="normal"); self.t4_ok.delete(0,"end"); self.t4_ok.insert(0, res); self.t4_ok.configure(state="readonly")
    except Exception: pass

# ---------- Tab 5 (Voltage Drop) ----------
def _build_tab5(self):
    f = self.tab5
    pad = {"padx":10, "pady":4}
    ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=12, sticky="w", **pad)
    self.t5_title = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad); self.t5_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
    self.t5_number = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Number", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad); self.t5_number.grid(row=1, column=4, sticky="we", **pad)
    self.t5_location = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad); self.t5_location.grid(row=1, column=6, sticky="we", **pad)
    self.t5_docno = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad); self.t5_docno.grid(row=2, column=1, sticky="we", **pad)
    self.t5_rev = ttk.Entry(f, width=10, style="App.TEntry", state="readonly"); ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad); self.t5_rev.grid(row=2, column=4, sticky="we", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=12, sticky="we", **pad)
    ttk.Label(f, text="Voltage Drop", style="App.Section.TLabel").grid(row=4, column=0, columnspan=12, sticky="w", **pad)

    ttk.Label(f, text="Calculation Number", style="App.TLabel").grid(row=5, column=0, sticky="w", **pad); self.t5_calcno = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); self.t5_calcno.grid(row=5, column=1, sticky="w", **pad)
    ttk.Label(f, text="Cable resistance r (Ω/km)", style="App.TLabel").grid(row=6, column=0, sticky="w", **pad); self.t5_r = ttk.Entry(f, width=14, style="App.TEntry"); self.t5_r.grid(row=6, column=1, sticky="w", **pad); self.t5_r.bind("<KeyRelease>", lambda e: (self._recalc_vdrop(), self._schedule_autosave()))
    ttk.Label(f, text="Cable reactance x (Ω/km)", style="App.TLabel").grid(row=6, column=3, sticky="w", **pad); self.t5_x = ttk.Entry(f, width=14, style="App.TEntry"); self.t5_x.grid(row=6, column=4, sticky="w", **pad); self.t5_x.bind("<KeyRelease>", lambda e: (self._recalc_vdrop(), self._schedule_autosave()))

    ttk.Label(f, text="Steady-state limit (%)", style="App.TLabel").grid(row=7, column=0, sticky="w", **pad); self.t5_ss_limit = ttk.Entry(f, width=10, style="App.TEntry"); self.t5_ss_limit.insert(0,"5.0"); self.t5_ss_limit.grid(row=7, column=1, sticky="w", **pad); self.t5_ss_limit.bind("<KeyRelease>", lambda e: (self._recalc_vdrop(), self._schedule_autosave()))
    ttk.Label(f, text="Starting limit (%)", style="App.TLabel").grid(row=7, column=3, sticky="w", **pad); self.t5_st_limit = ttk.Entry(f, width=10, style="App.TEntry"); self.t5_st_limit.insert(0,"15.0"); self.t5_st_limit.grid(row=7, column=4, sticky="w", **pad); self.t5_st_limit.bind("<KeyRelease>", lambda e: (self._recalc_vdrop(), self._schedule_autosave()))

    ttk.Label(f, text="Steady VD (V)", style="App.TLabel").grid(row=8, column=0, sticky="w", **pad); self.t5_ss_v = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t5_ss_v.grid(row=8, column=1, sticky="w", **pad)
    ttk.Label(f, text="Steady VD (%)", style="App.TLabel").grid(row=8, column=3, sticky="w", **pad); self.t5_ss_pct = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t5_ss_pct.grid(row=8, column=4, sticky="w", **pad)

    ttk.Label(f, text="Starting VD (V)", style="App.TLabel").grid(row=9, column=0, sticky="w", **pad); self.t5_st_v = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t5_st_v.grid(row=9, column=1, sticky="w", **pad)
    ttk.Label(f, text="Starting VD (%)", style="App.TLabel").grid(row=9, column=3, sticky="w", **pad); self.t5_st_pct = ttk.Entry(f, width=14, style="App.TEntry", state="readonly"); self.t5_st_pct.grid(row=9, column=4, sticky="w", **pad)

    ttk.Label(f, text="Steady OK?", style="App.TLabel").grid(row=10, column=0, sticky="w", **pad); self.t5_ss_ok = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); self.t5_ss_ok.grid(row=10, column=1, sticky="w", **pad)
    ttk.Label(f, text="Starting OK?", style="App.TLabel").grid(row=10, column=3, sticky="w", **pad); self.t5_st_ok = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); self.t5_st_ok.grid(row=10, column=4, sticky="w", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=11, column=0, columnspan=12, sticky="we", padx=8, pady=8)
    ttk.Button(f, text="Save/Update Calculation", style="App.TButton", command=self.on_save_calc).grid(row=12, column=0, sticky="w", **pad)
    self.btn5_prev = ttk.Button(f, text="Prev", style="App.TButton", command=self.on_prev_calc); self.btn5_prev.grid(row=12, column=6, sticky="e", **pad)
    self.btn5_next = ttk.Button(f, text="Next", style="App.TButton", command=self.on_next_calc); self.btn5_next.grid(row=12, column=7, sticky="e", **pad)

    for c in (1,4,6): f.grid_columnconfigure(c, weight=1)

def _recalc_vdrop(self):
    try:
        phase = (self.t2_phase.get() or "3").strip()
        V_ll = float(self.t2_vll.get() or 0.0)
        V_ph = calc_phase_voltage(V_ll, phase)
        L_km = float(self.t2_len.get() or 0.0) / 1000.0
        Ib = float(self.t2_flc.get() or 0.0)
        pf = float(self.t2_pf.get() or 0.0)
        spf = float(self.t2_spf.get() or 0.0)
        Ist = float(self.t2_istart.get() or 0.0)
        r = float(self.t5_r.get() or 0.0)
        x = float(self.t5_x.get() or 0.0)
        ss_lim = float(self.t5_ss_limit.get() or 100.0)
        st_lim = float(self.t5_st_limit.get() or 100.0)
    except Exception:
        return
    try:
        ang = math.acos(max(0.0, min(1.0, pf))) if pf>0 else 0.0
        sinp = math.sin(ang); cosp = pf
        if phase == "3":
            vd_ss = math.sqrt(3.0) * Ib * (r*cosp + x*sinp) * L_km
            vd_pct = (vd_ss / V_ll * 100.0) if V_ll>0 else 0.0
        else:
            vd_ss = 2.0 * Ib * (r*cosp + x*sinp) * L_km
            vd_pct = (vd_ss / V_ph * 100.0) if V_ph>0 else 0.0
        self.t5_ss_v.configure(state="normal"); self.t5_ss_v.delete(0,"end"); self.t5_ss_v.insert(0, f"{vd_ss:.2f}"); self.t5_ss_v.configure(state="readonly")
        self.t5_ss_pct.configure(state="normal"); self.t5_ss_pct.delete(0,"end"); self.t5_ss_pct.insert(0, f"{vd_pct:.2f}"); self.t5_ss_pct.configure(state="readonly")
        ss_ok = "OK" if vd_pct <= ss_lim else "FAIL"
        self.t5_ss_ok.configure(state="normal"); self.t5_ss_ok.delete(0,"end"); self.t5_ss_ok.insert(0, ss_ok); self.t5_ss_ok.configure(state="readonly")
    except Exception:
        pass
    try:
        angs = math.acos(max(0.0, min(1.0, spf))) if spf>0 else 0.0
        sins = math.sin(angs); coss = spf
        if phase == "3":
            vd_st = math.sqrt(3.0) * Ist * (r*coss + x*sins) * L_km
            vdsp = (vd_st / V_ll * 100.0) if V_ll>0 else 0.0
        else:
            vd_st = 2.0 * Ist * (r*coss + x*sins) * L_km
            vdsp = (vd_st / V_ph * 100.0) if V_ph>0 else 0.0
        self.t5_st_v.configure(state="normal"); self.t5_st_v.delete(0,"end"); self.t5_st_v.insert(0, f"{vd_st:.2f}"); self.t5_st_v.configure(state="readonly")
        self.t5_st_pct.configure(state="normal"); self.t5_st_pct.delete(0,"end"); self.t5_st_pct.insert(0, f"{vdsp:.2f}"); self.t5_st_pct.configure(state="readonly")
        st_ok = "OK" if vdsp <= st_lim else "FAIL"
        self.t5_st_ok.configure(state="normal"); self.t5_st_ok.delete(0,"end"); self.t5_st_ok.insert(0, st_ok); self.t5_st_ok.configure(state="readonly")
    except Exception:
        pass

# ---------- Tab 6 (Summary & Export) ----------
def _build_tab6(self):
    f = self.tab6
    pad = {"padx":10, "pady":4}
    ttk.Label(f, text="Project Information", style="App.Section.TLabel").grid(row=0, column=0, columnspan=12, sticky="w", **pad)
    self.t6_title = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Title", style="App.TLabel").grid(row=1, column=0, sticky="w", **pad); self.t6_title.grid(row=1, column=1, columnspan=2, sticky="we", **pad)
    self.t6_number = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Project Number", style="App.TLabel").grid(row=1, column=3, sticky="w", **pad); self.t6_number.grid(row=1, column=4, sticky="we", **pad)
    self.t6_location = ttk.Entry(f, width=25, style="App.TEntry", state="readonly"); ttk.Label(f, text="Location", style="App.TLabel").grid(row=1, column=5, sticky="w", **pad); self.t6_location.grid(row=1, column=6, sticky="we", **pad)
    self.t6_docno = ttk.Entry(f, width=50, style="App.TEntry", state="readonly"); ttk.Label(f, text="Document Number", style="App.TLabel").grid(row=2, column=0, sticky="w", **pad); self.t6_docno.grid(row=2, column=1, sticky="we", **pad)
    self.t6_rev = ttk.Entry(f, width=10, style="App.TEntry", state="readonly"); ttk.Label(f, text="Rev No", style="App.TLabel").grid(row=2, column=3, sticky="w", **pad); self.t6_rev.grid(row=2, column=4, sticky="we", **pad)

    ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=12, sticky="we", padx=8, pady=6)
    ttk.Label(f, text="Summary", style="App.Section.TLabel").grid(row=4, column=0, columnspan=12, sticky="w", **pad)
    self.t6_calcno = ttk.Entry(f, width=12, style="App.TEntry", state="readonly"); ttk.Label(f, text="Calculation Number", style="App.TLabel").grid(row=5, column=0, sticky="w", **pad); self.t6_calcno.grid(row=5, column=1, sticky="w", **pad)

    self.t6_summary = tk.Text(f, width=120, height=20, bg=self.EBG); self.t6_summary.grid(row=6, column=0, columnspan=12, sticky="nsew", **pad)
    ttk.Button(f, text="Refresh", style="App.TButton", command=self._refresh_summary).grid(row=7, column=0, sticky="w", **pad)
    ttk.Button(f, text="Export current calc to CSV", style="App.TButton", command=self.on_export_csv).grid(row=7, column=1, sticky="w", **pad)
    ttk.Button(f, text="Open project folder", style="App.TButton", command=self.on_open_project_folder).grid(row=7, column=2, sticky="w", **pad)
    for c in (1,4,6,8,10): f.grid_columnconfigure(c, weight=1)

def _refresh_summary(self):
    def g(e):
        try: return e.get().strip()
        except Exception: return ""
    lines = []
    lines.append(f"Calc No: {g(self.t2_calcno)}  |  Circuit: {g(self.t2_circuit)}  |  Cable Tag: {g(self.t2_cable)}")
    lines.append(f"From: {g(self.t2_from_desc)} ({g(self.t2_from_tag)})  →  To: {g(self.t2_to_desc)} ({g(self.t2_to_tag)})")
    lines.append("")
    lines.append("System")
    lines.append(f"  Phase: {g(self.t2_phase)}  VLL: {g(self.t2_vll)} V  Vph: {g(self.t2_vph)} V  kVA: {g(self.t2_kva)}  FLC: {g(self.t2_flc)} A  Istart: {g(self.t2_istart)} A")
    lines.append(f"  PF: {g(self.t2_pf)}  Eff%: {g(self.t2_eff)}  StartPF: {g(self.t2_spf)}  Length: {g(self.t2_len)} m")
    lines.append("")
    lines.append("Protection")
    lines.append(f"  Device In: {g(self.t2_pd)} A  Overload%: {g(self.t2_ol)}  Type: {g(self.t4_ptype)}  I2×In: {g(self.t4_i2)}  Result: {g(self.t4_ok)}")
    lines.append("")
    lines.append("Installation & Environment")
    lines.append(f"  Cable: {g(self.t3_type_cable)}  Install: {g(self.t3_type_install)}  Core: {g(self.t3_core_type)} {g(self.t3_formation)}")
    lines.append(f"  Ground/Ducts CF: Depth {g(self.t3_depth_cf)}, Soil {g(self.t3_soil_cf)}, Group {g(self.t3_ground_cf)}; Free Air CF: {g(self.t3_air_df)}; Temp CF: {g(self.t3_temp_df)}")
    lines.append(f"  ODF: {g(self.t3_odf)}")
    lines.append("")
    lines.append("Voltage Drop")
    lines.append(f"  r: {g(self.t5_r)} Ω/km  x: {g(self.t5_x)} Ω/km  |  Steady: {g(self.t5_ss_v)} V ({g(self.t5_ss_pct)}%), {g(self.t5_ss_ok)}  |  Start: {g(self.t5_st_v)} V ({g(self.t5_st_pct)}%), {g(self.t5_st_ok)}")
    self.t6_summary.delete("1.0","end"); self.t6_summary.insert("1.0", "\n".join(lines))

def on_export_csv(self):
    import csv
    if not self.store:
        messagebox.showwarning("Export", "Save or open a project first.")
        return
    try:
        payload = self._collect_calc_dict()
        proj = self.store.read_project_info().get("Project Number", "Project")
        fname = f"{_safe_filename(proj)}_{_safe_filename(payload.get('CalculationNumber','row'))}.csv"
        out = os.path.join(PROJECT_DIR, fname)
        headers = list(payload.keys())
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader(); w.writerow(payload)
        messagebox.showinfo("Export", f"CSV written:\n{out}")
    except Exception as e:
        messagebox.showerror("Export", f"{type(e).__name__}: {e}")

def on_open_project_folder(self):
    try:
        folder = PROJECT_DIR
        if os.name == "nt":
            os.startfile(folder)  # type: ignore
        elif sys.platform == "darwin":
            import subprocess; subprocess.Popen(["open", folder])
        else:
            import subprocess; subprocess.Popen(["xdg-open", folder])
    except Exception as e:
        messagebox.showerror("Open Folder", f"{type(e).__name__}: {e}")

# ---------- Project load/save + search (Tab 1 buttons) ----------
def _load_project(self, path: str):
    try:
        if not HAVE_XLSX:
            messagebox.showerror("Excel backend", "openpyxl is not available."); return
        if not path: return
        self.store = ProjectStore(path); self.store.open_or_create()
        kv = self.store.read_project_info()
        self._fill_tab1_from_kv(kv)
        self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv); self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
        try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(path)}")
        except Exception: pass
        self._tab2_boot_loaded = False
        self._refresh_calc_order(); self._ensure_default_calc_loaded(); add_to_recents(path)
        messagebox.showinfo("Project", "Project opened."); self.nb.select(self.tab2)
    except Exception as e:
        messagebox.showerror("Open Project", f"{type(e).__name__}: {e}")

def on_create_project(self):
    try:
        kv = self._collect_project_info(); proj_no = (kv.get("Project Number") or "").strip()
        if not proj_no:
            messagebox.showwarning("Create Project", "Enter a Project Number on Tab 1 first."); return
        path = project_filepath(proj_no); self.store = ProjectStore(path); self.store.open_or_create(); self.store.write_project_info(kv)
        self._fill_tab1_from_kv(kv); add_to_recents(path)
        self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv); self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
        try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(path)}")
        except Exception: pass
        self._tab2_boot_loaded = False; self._refresh_calc_order(); self._ensure_default_calc_loaded()
        messagebox.showinfo("Create Project", f"Created: {os.path.basename(path)}"); self.nb.select(self.tab2)
    except Exception as e:
        messagebox.showerror("Create Project", f"{type(e).__name__}: {e}")

def on_open_recent(self):
    try:
        initial = PROJECT_DIR if os.path.isdir(PROJECT_DIR) else os.path.expanduser("~")
        path = filedialog.askopenfilename(title="Open Project Workbook", initialdir=initial, filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            recents = get_recents(); path = recents[0] if recents else ""
            if not path: return
        self._load_project(path)
    except Exception as e:
        messagebox.showerror("Open Project", f"{type(e).__name__}: {e}")

def on_save_project(self):
    try:
        kv = self._collect_project_info(); proj_no = (kv.get("Project Number") or "").strip()
        if not proj_no:
            messagebox.showwarning("Save Project", "Enter a Project Number on Tab 1 first."); return
        if not self.store:
            path = project_filepath(proj_no); self.store = ProjectStore(path); self.store.open_or_create()
        self.store.write_project_info(kv); self._fill_tab1_from_kv(kv); add_to_recents(self.store.path)
        self._mirror_to_tab2(kv); self._mirror_to_tab3_project(kv); self._mirror_to_tab4_project(kv); self._mirror_to_tab5_project(kv); self._mirror_to_tab6_project(kv)
        try: self.title(f"Cable Calculation — {APP_VERSION}  |  {os.path.basename(self.store.path)}")
        except Exception: pass
        messagebox.showinfo("Save Project", "Project information saved.")
    except Exception as e:
        messagebox.showerror("Save Project", f"{type(e).__name__}: {e}")

def on_search_project(self):
    try:
        q = (self.t1_search.get() or "").strip().lower()
        if not q:
            messagebox.showinfo("Search Project", "Type part of a filename into the search box first."); return
        pattern = os.path.join(PROJECT_DIR, "*.xlsx"); candidates = []
        try:
            for p in glob.glob(pattern):
                if q in os.path.basename(p).lower(): candidates.append(p)
        except Exception: pass
        if not candidates:
            messagebox.showinfo("Search Project", "No matching .xlsx found in the projects folder."); return
        self._load_project(candidates[0])
    except Exception as e:
        messagebox.showerror("Search Project", f"{type(e).__name__}: {e}")

# ---------- Close handler ----------
def on_close(self):
    try:
        if getattr(self, "_autosave_job", None):
            self.after_cancel(self._autosave_job); self._autosave_job = None
    except Exception: pass
    try:
        if self.dirty_calc and self.store:
            if messagebox.askyesno("Unsaved changes", "Save current calculation before closing?"):
                self.on_save_calc()
        if self.dirty_project and self.store:
            if messagebox.askyesno("Project changes", "Save project info before closing?"):
                self.on_save_project()
    except Exception: pass
    self.destroy()

# ---------- Main launcher ----------
def main():
    app = CableCalcApp(); app.mainloop()

if __name__ == "__main__":
    main()

